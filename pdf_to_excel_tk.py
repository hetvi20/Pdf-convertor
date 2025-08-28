import tkinter as tk
from tkinter import filedialog, messagebox
import pdfplumber
import pytesseract
from pdf2image import convert_from_path
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from PIL import Image

# ----------------------------
# PDF to Excel Converter Class

class PDFtoExcelApp:
    def __init__(self, root):
        self.root = root
        self.root.title("PDF to Excel Converter")
        self.root.geometry("400x250")

        # Upload button
        self.upload_btn = tk.Button(root, text="Upload PDF", command=self.upload_pdf, width=20, height=2)
        self.upload_btn.pack(pady=10)

        # Convert button
        self.convert_btn = tk.Button(root, text="Convert to Excel", command=self.convert_to_excel, width=20, height=2)
        self.convert_btn.pack(pady=10)

        # Exit button
        self.exit_btn = tk.Button(root, text="Exit", command=self.exit_app, width=20, height=2, bg="red", fg="white")
        self.exit_btn.pack(pady=10)

        self.pdf_path = None

    def upload_pdf(self):
        file_path = filedialog.askopenfilename(filetypes=[("PDF Files", "*.pdf")])
        if file_path:
            self.pdf_path = file_path
            messagebox.showinfo("File Selected", f"Selected File:\n{file_path}")

    def convert_to_excel(self):
        if not self.pdf_path:
            messagebox.showerror("Error", "Please select a PDF file first.")
            return

        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "PDF Data"

            row_idx = 1
            extracted = False

            # Extract with pdfplumber (best for digital PDFs)
            with pdfplumber.open(self.pdf_path) as pdf:
                for page_num, page in enumerate(pdf.pages[:50], start=1):  # handle up to 50 pages
                    tables = page.extract_tables()

                    if tables:
                        # Append all tables continuously
                        for table in tables:
                            for row in table:
                                ws.append([cell.strip() if cell else "" for cell in row])
                            row_idx = ws.max_row + 1
                        extracted = True
                    else:
                        # Append text line by line
                        text = page.extract_text()
                        if text:
                            for line in text.split("\n"):
                                ws.cell(row=row_idx, column=1, value=line.strip())
                                row_idx += 1
                            extracted = True

            # OCR fallback for scanned PDFs
            if not extracted:
                images = convert_from_path(self.pdf_path, dpi=300, first_page=1, last_page=50)
                for page_num, img in enumerate(images, start=1):
                    text = pytesseract.image_to_string(img)
                    for line in text.split("\n"):
                        if line.strip():
                            ws.cell(row=row_idx, column=1, value=line.strip())
                            row_idx += 1

            # Auto-adjust column widths for readability
            for col in ws.columns:
                max_length = 0
                col_letter = get_column_letter(col[0].column)
                for cell in col:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                ws.column_dimensions[col_letter].width = min(max_length + 2, 50)

            # Save Excel file
            save_path = filedialog.asksaveasfilename(defaultextension=".xlsx",
                                                     filetypes=[("Excel Files", "*.xlsx")])
            if save_path:
                wb.save(save_path)
                messagebox.showinfo("Success", f"PDF converted successfully!\nSaved as {save_path}")

        except Exception as e:
            messagebox.showerror("Error", f"Conversion failed:\n{e}")

    def exit_app(self):
        self.root.quit()


# ----------------------------
# Run Tkinter PDF Convertor App

if __name__ == "__main__":
    root = tk.Tk()
    app = PDFtoExcelApp(root)
    root.mainloop()
